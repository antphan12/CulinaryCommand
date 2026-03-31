"""
get_margin_edge_order_details.py

Reads an Excel file produced by get_prairie_canary_restaurant_orders.py,
fetches the line items for each order from the Margin Edge API,
and writes a new Excel sheet containing the vendorItemName for each line item.

Usage:
    python3 get_margin_edge_order_details.py --input-file <path>.xlsx --restaurant-id <restaurantUnitId>

Environment variables (set in .env):
    ME_API_KEY  - Margin Edge API key
"""
# AI-ASSISTED

import argparse
import os
import sys
import time
from datetime import datetime, timezone
from typing import Optional, Union

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
import requests
from dotenv import load_dotenv

load_dotenv()

MARGIN_EDGE_API_BASE_URL = os.getenv("ME_API_BASE_URL", "https://api.marginedge.com/public")
MARGIN_EDGE_MIN_REQUEST_INTERVAL = 0.5
_margin_edge_last_request_time: float = 0.0


# ---------------------------------------------------------------------------
# API helpers
# ---------------------------------------------------------------------------

def get_margin_edge_session() -> requests.Session:
    api_key = os.environ.get("ME_API_KEY")
    if not api_key:
        print("ERROR: ME_API_KEY environment variable is not set.", file=sys.stderr)
        sys.exit(1)
    session = requests.Session()
    session.headers.update({
        "X-API-KEY": api_key,
        "Accept": "application/json",
    })
    return session


def make_margin_edge_request(session: requests.Session, path: str, params: Optional[dict] = None, retries: int = 3) -> Union[list, dict]:
    global _margin_edge_last_request_time
    url = f"{MARGIN_EDGE_API_BASE_URL}/{path.lstrip('/')}"

    elapsed = time.monotonic() - _margin_edge_last_request_time
    if elapsed < MARGIN_EDGE_MIN_REQUEST_INTERVAL:
        time.sleep(MARGIN_EDGE_MIN_REQUEST_INTERVAL - elapsed)

    for attempt in range(1, retries + 1):
        _margin_edge_last_request_time = time.monotonic()
        response = session.get(url, params=params, timeout=30)

        if response.status_code == 429:
            retry_after = float(response.headers.get("Retry-After", 60))
            print(f"  Rate limited. Waiting {retry_after:.0f}s (attempt {attempt}/{retries})...")
            time.sleep(retry_after)
            continue

        response.raise_for_status()
        return response.json()

    raise RuntimeError(f"Rate limit exceeded after {retries} retries: {url}")


def fetch_order_line_items(session: requests.Session, order_id: str, restaurant_id: str) -> list[dict]:
    order_detail = make_margin_edge_request(session, f"orders/{order_id}", params={"restaurantUnitId": restaurant_id})
    if not isinstance(order_detail, dict):
        return []
    return order_detail.get("lineItems", [])


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def read_orders_from_excel(input_file: str) -> list[dict]:
    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook.active

    headers = [cell.value for cell in worksheet[1]]
    orders = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        if row_data.get("Order ID"):
            orders.append({
                "orderId": str(row_data.get("Order ID", "")),
                "createdDate": str(row_data.get("Created Date", "")),
                "vendorName": str(row_data.get("Vendor Name", "")),
            })
    return orders


def apply_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="2F5496")
    cell.alignment = Alignment(horizontal="center")


def apply_group_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(fill_type="solid", fgColor="4472C4")
    cell.alignment = Alignment(horizontal="left")


def export_line_items_to_excel(order_line_items: list[dict], restaurant_id: str) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Order Line Items"

    column_headers = ["Order ID", "Created Date", "Vendor Name", "Vendor Item Name"]
    for col_index, header in enumerate(column_headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        apply_header_style(cell)

    current_row = 2
    for order_entry in order_line_items:
        order_id = order_entry["orderId"]
        created_date = order_entry["createdDate"]
        vendor_name = order_entry["vendorName"]
        line_items = order_entry["lineItems"]

        if not line_items:
            worksheet.cell(row=current_row, column=1, value=order_id)
            worksheet.cell(row=current_row, column=2, value=created_date)
            worksheet.cell(row=current_row, column=3, value=vendor_name)
            worksheet.cell(row=current_row, column=4, value="(no line items)")
            current_row += 1
            continue

        for line_item in line_items:
            vendor_item_name = line_item.get("vendorItemName", "N/A")
            worksheet.cell(row=current_row, column=1, value=order_id)
            worksheet.cell(row=current_row, column=2, value=created_date)
            worksheet.cell(row=current_row, column=3, value=vendor_name)
            worksheet.cell(row=current_row, column=4, value=vendor_item_name)
            current_row += 1

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

    output_filename = f"order_line_items_{restaurant_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    workbook.save(output_filename)
    print(f"\nExported to {output_filename}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description="Fetch Margin Edge line items for each order in an Excel file")
    parser.add_argument("--input-file", type=str, required=True, help="Path to the Excel file from get_prairie_canary_restaurant_orders.py")
    parser.add_argument("--restaurant-id", type=str, required=True, help="Margin Edge restaurantUnitId")
    args = parser.parse_args()

    if not os.path.exists(args.input_file):
        print(f"ERROR: File not found: {args.input_file}", file=sys.stderr)
        sys.exit(1)

    print(f"Reading orders from {args.input_file}...")
    orders = read_orders_from_excel(args.input_file)

    if not orders:
        print("No orders found in the Excel file.")
        sys.exit(0)

    print(f"Found {len(orders)} order(s). Fetching line items...\n")

    margin_edge_session = get_margin_edge_session()
    order_line_items = []

    for index, order in enumerate(orders, start=1):
        order_id = order["orderId"]
        print(f"  [{index}/{len(orders)}] Fetching line items for order {order_id}...")
        try:
            line_items = fetch_order_line_items(margin_edge_session, order_id, args.restaurant_id)
            order_line_items.append({**order, "lineItems": line_items})
            print(f"    Found {len(line_items)} line item(s).")
        except requests.HTTPError as error:
            print(f"    Failed to fetch order {order_id}: {error}")
            order_line_items.append({**order, "lineItems": []})

    export_line_items_to_excel(order_line_items, args.restaurant_id)


if __name__ == "__main__":
    main()