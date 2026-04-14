"""
get_prairie_canary_restaurant_orders.py

Fetches orders and their line-item details from the Margin Edge API
for a given restaurant unit over the past month.

Usage:
    python3 get_prairie_canary_restaurant_orders.py --restaurant-id <restaurantUnitId>

Environment variables (set in .env):
    ME_API_KEY  - Margin Edge API key
"""
# AI-ASSISTED

import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from typing import Optional, Union

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests
from dotenv import load_dotenv

load_dotenv()

MARGIN_EDGE_API_BASE_URL = os.getenv("ME_API_BASE_URL", "https://api.marginedge.com/public")
MARGIN_EDGE_MIN_REQUEST_INTERVAL = 0.5
_margin_edge_last_request_time: float = 0.0


def get_margin_edge_session() -> requests.Session:
    api_key = os.environ.get("ME_API_KEY")
    if not api_key:
        print("ERROR: ME_API_KEY environment variable is not set.", file=sys.stderr)
        sys.exit(1)
    session = requests.Session()
    session.headers.update({
        "X-API-KEY": api_key,
        "Accept": "application/json",
    })
    return session


def make_margin_edge_request(session: requests.Session, path: str, params: Optional[dict] = None, retries: int = 3) -> Union[list, dict]:
    global _margin_edge_last_request_time
    url = f"{MARGIN_EDGE_API_BASE_URL}/{path.lstrip('/')}"

    elapsed = time.monotonic() - _margin_edge_last_request_time
    if elapsed < MARGIN_EDGE_MIN_REQUEST_INTERVAL:
        time.sleep(MARGIN_EDGE_MIN_REQUEST_INTERVAL - elapsed)

    for attempt in range(1, retries + 1):
        _margin_edge_last_request_time = time.monotonic()
        response = session.get(url, params=params, timeout=30)

        if response.status_code == 429:
            retry_after = float(response.headers.get("Retry-After", 60))
            print(f"  Rate limited. Waiting {retry_after:.0f}s (attempt {attempt}/{retries})...")
            time.sleep(retry_after)
            continue

        response.raise_for_status()
        return response.json()

    raise RuntimeError(f"Rate limit exceeded after {retries} retries: {url}")


def unwrap_margin_edge_response(api_response: Union[list, dict]) -> list[dict]:
    if isinstance(api_response, list):
        return api_response
    for key in ("orders", "content", "data", "items", "results"):
        if key in api_response and isinstance(api_response[key], list):
            return api_response[key]
    return [api_response] if api_response else []


def fetch_orders(session: requests.Session, restaurant_id: str, start_date: str, end_date: str) -> list[dict]:
    all_orders = []
    next_page_cursor = None

    while True:
        params = {
            "restaurantUnitId": restaurant_id,
            "startDate": start_date,
            "endDate": end_date,
        }
        if next_page_cursor:
            params["nextPage"] = next_page_cursor

        api_response = make_margin_edge_request(session, "orders", params=params)
        all_orders.extend(unwrap_margin_edge_response(api_response))

        next_page_cursor = api_response.get("nextPage") if isinstance(api_response, dict) else None
        if not next_page_cursor:
            break

        print(f"  Fetching next page...")

    return all_orders


def fetch_order_detail(session: requests.Session, order_id: str) -> dict:
    api_response = make_margin_edge_request(session, f"orders/{order_id}")
    return api_response if isinstance(api_response, dict) else {}


def print_order_summary(order: dict) -> None:
    order_id    = order.get("orderId", "N/A")
    created_date  = order.get("createdDate", "N/A")
    vendor_name = order.get("vendorName", "N/A")
    print(f"  Order {order_id} | {created_date} | {vendor_name}")


def print_order_detail(order_detail: dict) -> None:
    line_items = order_detail.get("lineItems") or order_detail.get("items") or order_detail.get("lines") or []
    if not line_items:
        print("    (no line items found)")
        return
    print(f"    {'Product':<40} {'Qty':>8} {'Unit':<10} {'Unit Price':>12} {'Total':>12}")
    print(f"    {'-'*40} {'-'*8} {'-'*10} {'-'*12} {'-'*12}")
    for line_item in line_items:
        product_name = line_item.get("productName") or line_item.get("name", "Unknown")
        quantity     = line_item.get("quantity") or line_item.get("qty", 0)
        unit         = line_item.get("unit") or line_item.get("purchaseUnit", "")
        unit_price   = line_item.get("unitPrice") or line_item.get("price", 0)
        line_total   = line_item.get("totalPrice") or line_item.get("extendedPrice") or (float(quantity or 0) * float(unit_price or 0))
        print(f"    {str(product_name):<40} {str(quantity):>8} {str(unit):<10} ${float(unit_price or 0):>11.2f} ${float(line_total or 0):>11.2f}")


def export_orders_to_excel(trimmed_orders: list[dict], restaurant_id: str) -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center")

    headers = ["Order ID", "Created Date", "Vendor Name"]
    for col_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=col_index, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row_index, order in enumerate(trimmed_orders, start=2):
        worksheet.cell(row=row_index, column=1, value=order.get("orderId"))
        worksheet.cell(row=row_index, column=2, value=order.get("createdDate"))
        worksheet.cell(row=row_index, column=3, value=order.get("vendorName"))

    for column in worksheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[column[0].column_letter].width = max_length + 4

    output_filename = f"orders_{restaurant_id}_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    workbook.save(output_filename)
    print(f"\nExported {len(trimmed_orders)} order(s) to {output_filename}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Fetch Margin Edge order details for a restaurant")
    parser.add_argument("--restaurant-id", type=str, required=True, help="Margin Edge restaurantUnitId")
    parser.add_argument("--json", action="store_true", help="Output raw JSON instead of formatted table")

    args = parser.parse_args()

    end_date   = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    start_date = (datetime.now(timezone.utc) - timedelta(days=60)).strftime("%Y-%m-%d")

    margin_edge_session = get_margin_edge_session()

    print(f"Fetching orders for restaurant {args.restaurant_id} ({start_date} → {end_date})...")

    orders = fetch_orders(margin_edge_session, args.restaurant_id, start_date, end_date)

    if not orders:
        print("No orders found for this date range.")
        sys.exit(0)

    print(f"Found {len(orders)} order(s).\n")
    print(f"DEBUG - available fields in first order: {list(orders[0].keys())}\n")

    trimmed_orders = [
        {
            "orderId": order.get("orderId", "N/A"),
            "createdDate": order.get("createdDate", "N/A"),
            "vendorName": order.get("vendorName", "N/A"),
        }
        for order in orders
    ]

    if args.json:
        print(json.dumps(trimmed_orders, indent=2))
        return

    for trimmed_order in trimmed_orders:
        print_order_summary(trimmed_order)

    export_orders_to_excel(trimmed_orders, args.restaurant_id)


if __name__ == "__main__":
    main()